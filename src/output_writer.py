"""Output writer — produces the processed Excel workbook and audit log."""

import logging
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from version import AGENT_NAME, AGENT_VERSION

log = logging.getLogger(__name__)

# ── Colour constants ───────────────────────────────────────────────────────────
NAVY       = "1F3864"
WHITE      = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
GREEN      = "E2EFDA"
YELLOW     = "FFF2CC"
ORANGE     = "FCE4D6"
RED        = "FF0000"
GRAY       = "F2F2F2"

STATUS_COLORS = {
    "Approved – Auto":        GREEN,
    "Pending – Manager Review": YELLOW,
    "Pending – VP Review":    ORANGE,
    "Validation Error":       RED,
    "Closed":                 GRAY,
    "Received":               GREEN,
    "Partially Received":     YELLOW,
}

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CURRENCY_COLS = {"Unit Price ($)", "Subtotal ($)", "Tax Amount ($)", "Total Amount ($)"}


def _header_cell(cell, text: str) -> None:
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _data_cell(cell, value: Any, col_name: str, status: str, row_idx: int) -> None:
    cell.value  = value
    cell.font   = Font(name="Arial", size=9,
                       color=WHITE if status == "Validation Error" else "000000")
    cell.border = BORDER

    bg = STATUS_COLORS.get(status, WHITE if row_idx % 2 != 0 else LIGHT_BLUE)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="right" if col_name in CURRENCY_COLS else "left",
        vertical="center",
    )
    if col_name in CURRENCY_COLS:
        cell.number_format = "#,##0.00"


def _write_register_sheet(ws, records: list[dict], title: str) -> None:
    if not records:
        ws.append(["No records"])
        return

    headers = list(records[0].keys())
    for col_idx, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=col_idx), h)
    ws.row_dimensions[1].height = 30

    for row_idx, rec in enumerate(records, 2):
        status = str(rec.get("Status", ""))
        for col_idx, h in enumerate(headers, 1):
            _data_cell(ws.cell(row=row_idx, column=col_idx), rec.get(h), h, status, row_idx)

    # Column widths
    for col_idx, h in enumerate(headers, 1):
        width = max(len(h) + 2, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_summary_sheet(ws, records: list[dict], thresholds: dict, run_ts: datetime) -> None:
    valid = [r for r in records if r.get("Status") != "Validation Error"]

    def to_dec(v):
        try:
            return Decimal(str(v).replace(",", "").replace("$", "").strip())
        except Exception:
            return Decimal("0")

    total_spend = sum(to_dec(r.get("Total Amount ($)", 0)) for r in valid)
    tier_counts: dict[str, int] = {}
    tier_spend:  dict[str, Decimal] = {}
    for r in valid:
        t = str(r.get("Approval Tier", "N/A"))
        tier_counts[t] = tier_counts.get(t, 0) + 1
        tier_spend[t]  = tier_spend.get(t, Decimal("0")) + to_dec(r.get("Total Amount ($)", 0))

    rows = [
        ("Run Timestamp",             run_ts.strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Agent Version",             AGENT_VERSION),
        ("Input Rows (Total)",        len(records)),
        ("Valid POs",                 len(valid)),
        ("Validation Errors",         len(records) - len(valid)),
        ("Total Spend ($)",           f"${total_spend:,.2f}"),
        ("Average PO Value ($)",      f"${total_spend / len(valid):,.2f}" if valid else "$0.00"),
        ("Auto-Approved POs",         tier_counts.get("Auto-Approve", 0)),
        ("Manager Approval POs",      tier_counts.get("Manager Approval", 0)),
        ("VP Approval POs",           tier_counts.get("VP Approval", 0)),
        ("Auto-Approve Spend ($)",    f"${tier_spend.get('Auto-Approve', 0):,.2f}"),
        ("Manager Approval Spend ($)",f"${tier_spend.get('Manager Approval', 0):,.2f}"),
        ("VP Approval Spend ($)",     f"${tier_spend.get('VP Approval', 0):,.2f}"),
        ("Manager Threshold",         f"${thresholds['manager']:,.2f}"),
        ("VP Threshold",              f"${thresholds['vp']:,.2f}"),
    ]

    for col_idx, h in enumerate(["Metric", "Value"], 1):
        _header_cell(ws.cell(row=1, column=col_idx), h)

    for row_idx, (metric, value) in enumerate(rows, 2):
        c1 = ws.cell(row=row_idx, column=1, value=metric)
        c2 = ws.cell(row=row_idx, column=2, value=value)
        for c in (c1, c2):
            c.font   = Font(name="Arial", size=10)
            c.border = BORDER
            c.fill   = PatternFill("solid", fgColor=LIGHT_BLUE if row_idx % 2 == 0 else WHITE)
        c2.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def write_outputs(
    records: list[dict[str, Any]],
    anomalies: list[str],
    out_dir: Path,
    run_ts: datetime,
    thresholds: dict,
    input_filename: str,
) -> tuple[Path, Path]:
    """Write the processed Excel workbook and audit log. Returns (xlsx_path, log_path)."""
    ts_str = run_ts.strftime("%Y%m%d_%H%M%S")

    # ── Excel workbook ──────────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1 — Full register
    ws1 = wb.active
    ws1.title = "PO Register"
    _write_register_sheet(ws1, records, "PO Register")

    # Sheet 2 — Approval queue (pending only)
    pending = [r for r in records if "Pending" in str(r.get("Status", ""))]
    ws2 = wb.create_sheet("Approval Queue")
    _write_register_sheet(ws2, pending, "Approval Queue")

    # Sheet 3 — Validation errors
    errors = [r for r in records if r.get("Status") == "Validation Error"]
    ws3 = wb.create_sheet("Validation Errors")
    _write_register_sheet(ws3, errors, "Validation Errors")

    # Sheet 4 — Run summary
    ws4 = wb.create_sheet("Run Summary")
    _write_summary_sheet(ws4, records, thresholds, run_ts)

    xlsx_path = out_dir / f"PO_Processed_{ts_str}.xlsx"
    wb.save(xlsx_path)
    log.info(f"Excel written → {xlsx_path}")

    # ── Audit log ───────────────────────────────────────────────────────────────
    valid = [r for r in records if r.get("Status") != "Validation Error"]
    error_records = [r for r in records if r.get("Status") == "Validation Error"]

    tier_counts: dict[str, int] = {}
    for r in valid:
        t = str(r.get("Approval Tier", "N/A"))
        tier_counts[t] = tier_counts.get(t, 0) + 1

    lines = [
        f"{AGENT_NAME} v{AGENT_VERSION} — Audit Log",
        f"Run Timestamp : {run_ts.isoformat()}",
        f"Input File    : {input_filename}",
        f"Total Rows    : {len(records)}",
        f"Valid POs     : {len(valid)}",
        f"Errors        : {len(error_records)}",
        "",
        "--- Approval Tier Distribution ---",
        *[f"  {tier}: {count}" for tier, count in tier_counts.items()],
        "",
        "--- Validation Errors ---",
        *(
            [f"  PO {r.get('PO Number', '?')}: {r.get('Validation Notes', '')}" for r in error_records]
            or ["  None"]
        ),
        "",
        "--- Anomalies ---",
        *(["  " + a for a in anomalies] or ["  None detected"]),
        "",
        "--- End of Log ---",
    ]

    log_path = out_dir / f"agent_log_{ts_str}.txt"
    log_path.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"Audit log written → {log_path}")

    return xlsx_path, log_path
